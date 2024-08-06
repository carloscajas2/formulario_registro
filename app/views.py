from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.http import HttpResponse
from .models import Agencia, Registro
import csv
from io import StringIO
from django.db import connection
import openpyxl
from openpyxl.utils import get_column_letter

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('manage')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    return render(request, 'login.html')


def manage_view(request):
    if not request.user.is_authenticated:
        return redirect('login')

    user_region_mapping = {
        'centro': 'CENTRO',
        'occidente': 'OCCIDENTE',
        'oriente': 'ORIENTE',
    }

    username = request.user.username
    if username in user_region_mapping:
        agencias = Agencia.objects.filter(region=user_region_mapping[username])
    else:
        agencias = Agencia.objects.all()

    agencias_datos = []
    for agencia in agencias:
        data = {
            'agencia': agencia,
            'ventanilla': {
                'contratados': 0, 'conectados': 0, 'nuevos': 0, 'bajas_medicas': 0, 'remplazo_plataforma_rac': 0,
                'externas_y_autobancos': 0, 'promotor_offline': 0
            },
            'plataforma': {
                'contratados': 0, 'conectados': 0, 'nuevos': 0, 'bajas_medicas': 0, 'remplazo_plataforma_rac': 0,
                'externas_y_autobancos': 0, 'promotor_offline': 0
            }
        }
        ventanilla = Registro.objects.filter(agencia=agencia.nom_age, area='Ventanilla').order_by('-timestamp').first()
        plataforma = Registro.objects.filter(agencia=agencia.nom_age, area='Plataforma').order_by('-timestamp').first()

        if ventanilla:
            data['ventanilla'].update({
                'contratados': ventanilla.contratados,
                'conectados': ventanilla.conectados,
                'nuevos': ventanilla.nuevos,
                'bajas_medicas': ventanilla.bajas_medicas,
                'remplazo_plataforma_rac': ventanilla.remplazo_plataforma_rac,
                'externas_y_autobancos': ventanilla.externas_y_autobancos,
                'promotor_offline': ventanilla.promotor_offline
            })

        if plataforma:
            data['plataforma'].update({
                'contratados': plataforma.contratados,
                'conectados': plataforma.conectados,
                'nuevos': plataforma.nuevos,
                'bajas_medicas': plataforma.bajas_medicas,
                'remplazo_plataforma_rac': plataforma.remplazo_plataforma_rac,
                'externas_y_autobancos': plataforma.externas_y_autobancos,
                'promotor_offline': plataforma.promotor_offline
            })

        agencias_datos.append(data)

    if request.method == 'POST':
        for data in agencias_datos:
            Registro.objects.create(
                agencia=data['agencia'].nom_age,
                area='Ventanilla',
                contratados=request.POST.get(f'contratados_vent_{data["agencia"].id_age}', 0),
                conectados=request.POST.get(f'conectados_vent_{data["agencia"].id_age}', 0),
                nuevos=request.POST.get(f'nuevos_vent_{data["agencia"].id_age}', 0),
                bajas_medicas=request.POST.get(f'bajas_medicas_vent_{data["agencia"].id_age}', 0),
                remplazo_plataforma_rac=request.POST.get(f'remplazo_plataforma_rac_vent_{data["agencia"].id_age}', 0),
                externas_y_autobancos=request.POST.get(f'externas_y_autobancos_vent_{data["agencia"].id_age}', 0),
                promotor_offline=request.POST.get(f'promotor_offline_vent_{data["agencia"].id_age}', 0)
            )
            Registro.objects.create(
                agencia=data['agencia'].nom_age,
                area='Plataforma',
                contratados=request.POST.get(f'contratados_plat_{data["agencia"].id_age}', 0),
                conectados=request.POST.get(f'conectados_plat_{data["agencia"].id_age}', 0),
                nuevos=request.POST.get(f'nuevos_plat_{data["agencia"].id_age}', 0),
                bajas_medicas=request.POST.get(f'bajas_medicas_plat_{data["agencia"].id_age}', 0),
                remplazo_plataforma_rac=request.POST.get(f'remplazo_plataforma_rac_plat_{data["agencia"].id_age}', 0),
                externas_y_autobancos=request.POST.get(f'externas_y_autobancos_plat_{data["agencia"].id_age}', 0),
                promotor_offline=request.POST.get(f'promotor_offline_plat_{data["agencia"].id_age}', 0)
            )
        messages.success(request, 'Registro guardado con éxito')

    return render(request, 'form.html', {'agencias_datos': agencias_datos})

def download_excel(request):
    # Definir la primera consulta SQL
    query1 = """
    with registros_ordenados as (
        select
            region,
            ciudad,
            agencia,
            strftime('%Y-%m-%d', datetime(timestamp, '-4 hours')) as fecha,
            area,
            contratados,
            conectados,
            bajas_medicas,
            nuevos,
            promotor_offline,
            remplazo_plataforma_rac,
            externas_y_autobancos,
            row_number() over (
                partition by region, area, agencia, strftime('%Y-%m-%d', datetime(timestamp, '-4 hours'))
                order by datetime(timestamp, '-4 hours') desc
            ) as rn
        from
            app_registro as a
            left join app_agencia as b on a.agencia = b.nom_age
    )
    select
        region,
        ciudad,
        agencia,
        fecha,
        area as canal,
        contratados,
        conectados,
        bajas_medicas,
        nuevos,
        promotor_offline,
        remplazo_plataforma_rac,
        externas_y_autobancos
    from
        registros_ordenados
    where
        rn = 1
    order by
        fecha,
        ciudad,
        agencia,
        area;
    """

    # Definir la segunda consulta SQL
    query2 = """
    with registros_ordenados as (
        select
            region,
            ciudad,
            agencia,
            strftime('%Y-%m-%d', datetime(timestamp, '-4 hours')) as fecha,
            area,
            contratados,
            conectados,
            bajas_medicas,
            nuevos,
            promotor_offline,
            remplazo_plataforma_rac,
            externas_y_autobancos,
            row_number() over (
                partition by region, area, agencia, strftime('%Y-%m-%d', datetime(timestamp, '-4 hours'))
                order by datetime(timestamp, '-4 hours') desc
            ) as rn
        from
            app_registro as a
            left join app_agencia as b on a.agencia = b.nom_age
    ),
    ultimos_registros as (
        select
            region,
            ciudad,
            agencia,
            fecha,
            area,
            contratados,
            conectados,
            bajas_medicas,
            nuevos,
            promotor_offline,
            remplazo_plataforma_rac,
            externas_y_autobancos
        from
            registros_ordenados
        where
            rn = 1
    )
    select
        region,
        area, 
        sum(contratados) as total_contratados,
        sum(conectados) as total_conectados,
        sum(bajas_medicas) as total_bajas_medicas,
        sum(nuevos) as total_nuevos,
        sum(promotor_offline) as total_promotor_offline,
        sum(remplazo_plataforma_rac) as total_remplazo_plataforma_rac,
        sum(externas_y_autobancos) as total_externas_y_autobancos
    from
        ultimos_registros
    group by
        region,
        area
    order by
        region,
        area;
    """

    # Ejecutar la primera consulta
    with connection.cursor() as cursor:
        cursor.execute(query1)
        rows1 = cursor.fetchall()

    # Ejecutar la segunda consulta
    with connection.cursor() as cursor:
        cursor.execute(query2)
        rows2 = cursor.fetchall()

    # Crear un libro de Excel
    workbook = openpyxl.Workbook()

    # Escribir la primera hoja
    sheet1 = workbook.active
    sheet1.title = 'Detalle por Agencia'
    headers1 = [
        'region', 'ciudad', 'agencia', 'fecha', 'canal', 'contratados',
        'conectados', 'bajas_medicas', 'nuevos', 'promotor_offline',
        'remplazo_plataforma_rac', 'externas_y_autobancos'
    ]
    for col_num, header in enumerate(headers1, 1):
        col_letter = get_column_letter(col_num)
        sheet1[f'{col_letter}1'] = header
    for row_num, row in enumerate(rows1, 2):
        for col_num, cell_value in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            sheet1[f'{col_letter}{row_num}'] = cell_value

    # Crear y escribir la segunda hoja
    sheet2 = workbook.create_sheet(title='Resumen por Región')
    headers2 = [
        'region', 'area', 'total_contratados', 'total_conectados',
        'total_bajas_medicas', 'total_nuevos', 'total_promotor_offline',
        'total_remplazo_plataforma_rac', 'total_externas_y_autobancos'
    ]
    for col_num, header in enumerate(headers2, 1):
        col_letter = get_column_letter(col_num)
        sheet2[f'{col_letter}1'] = header
    for row_num, row in enumerate(rows2, 2):
        for col_num, cell_value in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            sheet2[f'{col_letter}{row_num}'] = cell_value

    # Crear la respuesta Http para el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="FormularioEstados.xlsx"'
    workbook.save(response)

    return response
